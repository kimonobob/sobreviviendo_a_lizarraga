#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Genera public/data/alicorp.json a partir del Excel de EEFF separados de Alicorp
y valida que ningun monto contradiga los estados financieros originales.

Uso:
    python scripts/build_data.py

El JSON es la UNICA fuente de datos del dashboard. Para actualizar con un
nuevo anio: agrega la columna al Excel y vuelve a correr este script.
"""
import json
import re
import sys
import unicodedata
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT.parent / "Alicorp_EEFF_Separados_2010-2025.xlsx"
OUT = ROOT / "public" / "data" / "alicorp.json"


def slug(text):
    t = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode()
    t = t.lower()
    t = re.sub(r"[^a-z0-9]+", "_", t).strip("_")
    return t


def num(v):
    """Devuelve int, float o None (celda vacia)."""
    if v is None or v == "":
        return None
    if isinstance(v, float):
        if v.is_integer():
            return int(v)
        return round(v, 6)
    if isinstance(v, int):
        return v
    return None


def row_values(row, ncols):
    """Valores de las 16 columnas de anios (indices 1..ncols)."""
    return [num(row[i]) if i < len(row) else None for i in range(1, ncols + 1)]


def has_numbers(vals):
    return any(v is not None for v in vals)


# ---------------------------------------------------------------------------
# Carga
# ---------------------------------------------------------------------------
if not XLSX.exists():
    sys.exit(f"No se encontro el Excel: {XLSX}")

wb = openpyxl.load_workbook(XLSX, data_only=True)

# Anios: fila de encabezado "Cuenta | 2010 | 2011 | ..." en ESF
esf = wb["ESF"]
years = None
for row in esf.iter_rows(values_only=True):
    if row and row[0] == "Cuenta":
        years = [int(y) for y in row[1:] if y is not None]
        break
if not years:
    sys.exit("No se pudo leer la fila de anios en ESF.")
NY = len(years)
print(f"Anios detectados: {years[0]}-{years[-1]} ({NY})")

# ---------------------------------------------------------------------------
# ESF: bloques -> cuentas
# ---------------------------------------------------------------------------
# start_header (uppercase exacto) -> total_row (label que cierra el bloque)
BLOQUES_DEF = [
    ("activo_corriente",    "Activo Corriente",    "activo",     "ACTIVO CORRIENTE",    "total activo corriente"),
    ("activo_no_corriente", "Activo No Corriente", "activo",     "ACTIVO NO CORRIENTE", "total activo no corriente"),
    ("pasivo_corriente",    "Pasivo Corriente",    "pasivo",     "PASIVO CORRIENTE",    "total pasivo corriente"),
    ("pasivo_no_corriente", "Pasivo No Corriente", "pasivo",     "PASIVO NO CORRIENTE", "total pasivo no corriente"),
    ("patrimonio",          "Patrimonio",          "patrimonio", "PATRIMONIO",          "total patrimonio"),
]
start_map = {d[3]: d for d in BLOQUES_DEF}
total_map = {d[4]: d[0] for d in BLOQUES_DEF}

bloques = {d[0]: {"id": d[0], "label": d[1], "lado": d[2], "total": None, "cuentas": []}
           for d in BLOQUES_DEF}

TOTALES_LABELS = {
    "total activo": "activo",
    "total pasivo": "pasivo",
    "total patrimonio": "patrimonio",
    "total pasivo y patrimonio": "pasivo_patrimonio",
}
totales = {}

current = None
for row in esf.iter_rows(values_only=True):
    label = row[0]
    if not isinstance(label, str) or not label.strip():
        continue
    lab = label.strip()
    low = lab.lower()

    if lab in start_map:            # encabezado de bloque
        current = start_map[lab][0]
        continue

    if low in total_map:            # fila total del bloque -> cierra
        bloques[total_map[low]]["total"] = row_values(row, NY)
        current = None
        continue

    if low in TOTALES_LABELS:       # totales globales (pueden repetir header+data)
        vals = row_values(row, NY)
        if has_numbers(vals):
            totales[TOTALES_LABELS[low]] = vals
        continue

    if current:                     # cuenta dentro de un bloque
        vals = row_values(row, NY)
        if has_numbers(vals):
            bloques[current]["cuentas"].append(
                {"id": slug(lab), "label": lab, "values": vals}
            )

# "TOTAL PATRIMONIO" cierra el bloque patrimonio (total_map) antes de llegar a
# TOTALES_LABELS, asi que el total global de patrimonio se toma del bloque.
if "patrimonio" not in totales and bloques["patrimonio"]["total"]:
    totales["patrimonio"] = bloques["patrimonio"]["total"]

esf_out = {
    "bloques": [bloques[d[0]] for d in BLOQUES_DEF],
    "totales": totales,
}

# ---------------------------------------------------------------------------
# ER: lineas canonicas (fieles al Excel, en orden)
# ---------------------------------------------------------------------------
ER_MAP = [
    ("Ventas a terceros",                                     "ventas_terceros"),
    ("Ventas a partes relacionadas",                          "ventas_relacionadas"),
    ("Total ingresos de actividades ordinarias",             "ingresos"),
    ("Costo de ventas",                                       "costo_ventas"),
    ("Utilidad bruta",                                        "utilidad_bruta"),
    ("Gastos de ventas y distribucion",                       "gastos_venta"),
    ("Gastos de administracion",                              "gastos_admin"),
    ("Resultado de operaciones con derivados de materias",    "derivados_mp"),
    ("Otros ingresos y gastos, neto",                         "otros_ingresos_gastos"),
    ("Utilidad operativa",                                    "utilidad_operativa"),
    ("Ingresos financieros",                                  "ingresos_financieros"),
    ("Gastos financieros",                                    "gastos_financieros"),
    ("Diferencias de cambio, neto",                           "diferencia_cambio"),
    ("Participacion en resultados netos de subsidiarias",     "part_subsidiarias"),
    ("Perdida neta por instrumentos financieros derivados",   "perdida_derivados"),
    ("Utilidad antes del impuesto a las ganancias",           "utilidad_antes_impuesto"),
    ("Gasto por impuesto a las ganancias",                    "impuesto"),
    ("Utilidad neta de operaciones continuas",                "utilidad_continuas"),
    ("Utilidad (perdida) de operaciones discontinuadas",      "discontinuadas"),
    ("Utilidad neta del ejercicio",                           "utilidad_neta"),
]
er_sheet = wb["ER"]
er_rows = {r[0].strip(): r for r in er_sheet.iter_rows(values_only=True)
           if isinstance(r[0], str) and r[0].strip()}

def find_er_row(prefix):
    for label, r in er_rows.items():
        if label.lower().startswith(prefix.lower()):
            return r
    return None

er_lineas = []
er_labels = {}
for prefix, rid in ER_MAP:
    r = find_er_row(prefix)
    if r is None:
        print(f"  [aviso] ER: no se encontro '{prefix}'")
        continue
    lab = r[0].strip()
    er_labels[rid] = lab
    er_lineas.append({"id": rid, "label": lab, "values": row_values(r, NY)})

er_out = {"lineas": er_lineas}

# ---------------------------------------------------------------------------
# Ratios (transportados; no se recalculan)
# ---------------------------------------------------------------------------
def ratio_format(label):
    low = label.lower()
    if "margen" in low or "roa" in low or "roe" in low:
        return "pct"          # el valor viene en decimal (0.28 = 28%)
    if "dias" in low or "ciclo" in low or "dio" in low or "dso" in low or "dpo" in low:
        return "dias"
    return "x"

# claves en forma normalizada (slug con espacios): sin acentos, parentesis ni barras
GROUP_HEADERS = {
    "liquidez y solvencia": ("liquidez", "Liquidez y solvencia"),
    "actividad eficiencia": ("actividad", "Actividad / eficiencia"),
    "rentabilidad margenes": ("margenes", "Margenes"),
    "rentabilidad retornos": ("retornos", "Retornos"),
    "descomposicion dupont": ("dupont", "DuPont"),
}
# ids limpios y estables para los ratios (por prefijo de etiqueta normalizada)
RATIO_IDS = [
    ("liquidez corriente",            "liquidez_corriente"),
    ("prueba acida",                  "prueba_acida"),
    ("deuda financiera",              "deuda_patrimonio"),
    ("pasivo total",                  "pasivo_patrimonio"),
    ("cobertura de intereses",        "cobertura_intereses"),
    ("rotacion de inventarios",       "rotacion_inventarios"),
    ("dias de inventario",            "dio"),
    ("dias de cobro",                 "dso"),
    ("dias de pago",                  "dpo"),
    ("ciclo de conversion",           "cce"),
    ("margen bruto",                  "margen_bruto"),
    ("margen operativo",              "margen_operativo"),
    ("margen neto",                   "margen_neto"),
    ("roa",                           "roa"),
    ("roe ut",                        "roe"),
    ("1 margen neto",                 "dp_margen"),
    ("2 rotacion de activos",         "dp_rotacion"),
    ("3 multiplicador",               "dp_multiplicador"),
    ("roe dupont",                    "roe_dupont"),
]

def ratio_id(label):
    norm = slug(label).replace("_", " ")
    for prefix, rid in RATIO_IDS:
        if norm.startswith(prefix):
            return rid
    return slug(label)

ratios_sheet = wb["Ratios"]
grupos = []
current_g = None
for row in ratios_sheet.iter_rows(values_only=True):
    label = row[0]
    if not isinstance(label, str) or not label.strip():
        continue
    lab = label.strip()
    key = slug(lab).replace("_", " ")
    if lab.lower().startswith("notas de calculo"):
        break
    if lab.startswith("•") or lab.startswith("•"):
        continue
    if key in GROUP_HEADERS:
        gid, glabel = GROUP_HEADERS[key]
        current_g = {"id": gid, "label": glabel, "items": []}
        grupos.append(current_g)
        continue
    vals = row_values(row, NY)
    if current_g is not None and has_numbers(vals):
        current_g["items"].append({
            "id": ratio_id(lab),
            "label": lab,
            "formato": ratio_format(lab),
            "values": vals,
        })

ratios_out = {"grupos": grupos}

# ---------------------------------------------------------------------------
# Meta
# ---------------------------------------------------------------------------
meta = {
    "empresa": "Alicorp S.A.A.",
    "base": "Separado / individual (NO consolidado)",
    "unidad": "S/ 000",
    "moneda": "PEN",
    "fuente": "SMV — EEFF separados auditados",
    "years": years,
    "notas": [
        "Base separada (matriz sola): el Activo y la Utilidad neta estan dominados por las inversiones en subsidiarias medidas por metodo de participacion. Interpretar como holding, no como desempeno operativo del grupo.",
        "Reduccion de capital 2024-2025 (capital emitido 847,192 → 686,226 → 569,573) reduce el patrimonio y eleva artificialmente ROE y apalancamiento de esos anios.",
        "Cambio de metodo en inversiones (costo → participacion) rompe la comparabilidad del patrimonio y del ROE entre 2011 y 2012.",
        "NIIF 16 (arrendamientos) desde 2018-2019 incrementa los pasivos financieros; Deuda/Patrimonio y cobertura de intereses no son estrictamente comparables antes vs. despues.",
        "2021 registra la unica perdida neta del periodo (participacion negativa en subsidiarias). 2014 tuvo utilidad minima por resultado de derivados de materias primas.",
        "Cifras en miles de soles (S/ 000). Cuadre verificado: Activo = Pasivo + Patrimonio en los 16 anios.",
    ],
}

data = {"meta": meta, "esf": esf_out, "er": er_out, "ratios": ratios_out}

# ---------------------------------------------------------------------------
# VALIDACION DE CUADRES (frente al Excel original)
# ---------------------------------------------------------------------------
print("\n== Validacion de cuadres ==")
errors = []


def approx(a, b, tol=1):
    if a is None or b is None:
        return a == b
    return abs(a - b) <= tol


def sum_cuentas(bloque, i):
    s = 0
    seen = False
    for c in bloque["cuentas"]:
        v = c["values"][i]
        if v is not None:
            s += v
            seen = True
    return s if seen else None


for i, y in enumerate(years):
    b = {bl["id"]: bl for bl in esf_out["bloques"]}
    # (a) subtotales de bloque == suma de sus cuentas
    for bid, bl in b.items():
        tot = bl["total"][i] if bl["total"] else None
        ssum = sum_cuentas(bl, i)
        if tot is not None and ssum is not None and not approx(tot, ssum):
            errors.append(f"{y}: {bid} total {tot} != suma cuentas {ssum}")
    # (b) Total activo == AC + ANC
    act = totales["activo"][i]
    ac = b["activo_corriente"]["total"][i]
    anc = b["activo_no_corriente"]["total"][i]
    if not approx(act, ac + anc):
        errors.append(f"{y}: Total activo {act} != AC+ANC {ac+anc}")
    # (c) Total pasivo == PC + PNC
    pas = totales["pasivo"][i]
    pc = b["pasivo_corriente"]["total"][i]
    pnc = b["pasivo_no_corriente"]["total"][i]
    if not approx(pas, pc + pnc):
        errors.append(f"{y}: Total pasivo {pas} != PC+PNC {pc+pnc}")
    # (d) Activo == Pasivo + Patrimonio
    pat = totales["patrimonio"][i]
    if not approx(act, pas + pat):
        errors.append(f"{y}: Activo {act} != Pasivo+Patrimonio {pas+pat}")
    # (e) ER: Ingresos - Costo == Utilidad bruta
    el = {l["id"]: l["values"] for l in er_lineas}
    ing, cv, ub = el["ingresos"][i], el["costo_ventas"][i], el["utilidad_bruta"][i]
    if not approx(ing + cv, ub):
        errors.append(f"{y}: Ingresos+Costo {ing+cv} != Utilidad bruta {ub}")
    # (f) ER cascada -> Utilidad operativa
    op_calc = ub + el["gastos_venta"][i] + el["gastos_admin"][i] \
        + (el["derivados_mp"][i] or 0) + (el["otros_ingresos_gastos"][i] or 0)
    if not approx(op_calc, el["utilidad_operativa"][i]):
        errors.append(f"{y}: cascada op {op_calc} != Utilidad operativa {el['utilidad_operativa'][i]}")
    # (g) ER -> Utilidad neta del ejercicio (reconciliacion completa)
    neta_calc = el["utilidad_operativa"][i] \
        + (el["ingresos_financieros"][i] or 0) + (el["gastos_financieros"][i] or 0) \
        + (el["diferencia_cambio"][i] or 0) + (el["part_subsidiarias"][i] or 0) \
        + (el["perdida_derivados"][i] or 0) + el["impuesto"][i] \
        + (el["discontinuadas"][i] or 0)
    if not approx(neta_calc, el["utilidad_neta"][i], tol=2):
        errors.append(f"{y}: cascada neta {neta_calc} != Utilidad neta {el['utilidad_neta'][i]}")

if errors:
    print(f"  [FALLO] {len(errors)} inconsistencias:")
    for e in errors:
        print("   -", e)
    sys.exit(1)
print(f"  [OK] {NY} anios cuadran: subtotales, Activo=Pasivo+Patrimonio y cascada del ER.")

# ---------------------------------------------------------------------------
# Escritura
# ---------------------------------------------------------------------------
OUT.parent.mkdir(parents=True, exist_ok=True)
with open(OUT, "w", encoding="utf-8") as f:
    json.dump(data, f, ensure_ascii=False, indent=1)
kb = OUT.stat().st_size / 1024
print(f"\nEscrito {OUT.relative_to(ROOT)} ({kb:.1f} KB)")
print(f"  bloques ESF: {len(esf_out['bloques'])} | lineas ER: {len(er_lineas)} | grupos ratios: {len(grupos)}")
